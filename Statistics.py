# coding:utf-8
import os
import json
import pymysql
from openpyxl import Workbook
from model.TorqueLog import TorqueLogFormat
from model.TorqueEvent import TorqueEventFormat
from sql.SqlConfigParser import SqlConfigParser
from sql.SqlConnector import SqlConnector

write_path_prefix = "./log_json"
root_path = './logs'
# root_path = 'E:/11 Python_WorkSpace/log'
excel_name = "./result_2019.xlsx"


# 获取文件夹下的文件列表
def get_file_list(folder_path):
    if os.path.isdir(folder_path):
        return os.listdir(folder_path)
    return 0


# 获取每行日志任务的状态
def get_job_status(line):
    tokens = line.split(";")
    if len(tokens) == 4:
        return tokens[1]
    return '0'


# 将每行日志转换成字典类型
def convert_to_dict(line):
    tokens = line.split(";")
    # print(tokens[3])
    if len(tokens[3]) > 0:
        split_tokens = tokens[3].split(" ")
        # print(split_tokens)
        log_dict = {}
        for token in split_tokens:
            tmp_tokens = token.split("=")
            # print(tmp_tokens)
            if len(tmp_tokens) == 2:
                # print(tmp_tokens[0] + " : ", tmp_tokens[1])
                log_dict[tmp_tokens[0]] = tmp_tokens[1]
            if len(tmp_tokens) == 3:
                log_dict[tmp_tokens[0]] = tmp_tokens[1] + tmp_tokens[2]
                print(tmp_tokens[2])
        return log_dict


# 将字典类型转换成json
def convert_to_json(job_list):
    return json.dumps(job_list, sort_keys=True)


# 写一个json文件至磁盘
def write_to_json_file(filename, src_list_file):
    with open(filename, "w") as file:
        json.dump(src_list_file, file, sort_keys=True)


# 读取文件列表
def read_file_list(root_path, file_list):
    torque_event_list = []
    for i in range(0, len(log_file_list)):
        log_file_path = os.path.join(root_path, file_list[i])
        if os.path.isfile(log_file_path):
            log_job_list = []
            with open(log_file_path, 'r') as log_file:
                for line in log_file:
                    # eachline = log_file.readline()
                    # print ("status: " + get_job_status(line) )
                    conn = SqlConnector()
                    torque_log_format = TorqueLogFormat(line)
                    sql = """
                    INSERT INTO t_log ( event_marker, job_num, event_detail )
                       VALUES ( %s, %s, %s);
                    """
                    # 执行SQL语句
                    conn.execute(sql, (
                        torque_log_format.event_marker, torque_log_format.job_num, torque_log_format.event_detail))

                    if 'E' == torque_log_format.event_marker:
                        torque_event_format = TorqueEventFormat(torque_log_format.event_detail)
                        # print(file_list[i] + " " + "Status: " + torque_log.event_marker + " needppn: " + str(torque_event_format.Resource_List_need_ppn) + " jobName: " + torque_event_format.job_name)
                        torque_event_list.append(torque_event_format)
                # 关闭连接
                # conn.close()
    return torque_event_list

    ######################################################################
    #         if 'E' == get_job_status(line):
    #             log_dict = convert_to_dict(line)
    #             # print(log_dict)
    #             log_job_list.append(log_dict)
    # # print(log_job_list)
    #
    # # list转换成json # 将json文件写入文件
    # job_json = convert_to_json(log_job_list)
    # # print(job_json)
    # write_json_file_path = os.path.join(write_path_prefix, file_list[i])
    # write_to_json_file(write_json_file_path + '.json', log_job_list)
    ######################################################################


# 获取类的自定义属性
def get_class_attr(clazz):
    attr_list = []
    for attr in dir(clazz):
        if "__" not in attr:
            attr_list.append(attr)
    return attr_list


# 创建一个Excel文件，并写入数据
def create_excel(filename, sheet_name, event_list):
    # 创建一个workbook 设置编码
    workbook = Workbook()
    # 创建一个worksheet
    # worksheet = workbook.create_sheet(sheet_name)
    worksheet = workbook.active
    # 写入标题行
    attr_list = get_class_attr(TorqueEventFormat)
    for i in range(len(attr_list)):
        # worksheet.write(0, i, attr_list[i])
        worksheet.cell(row=1, column=i + 1).value = attr_list[i]
    # workbook.save(excel_name)
    # 写入统计数据
    for i in range(len(event_list)):
        # print(attr_list)
        for j in range(len(attr_list)):
            # print(attr_list[j] + " value: " + str(getattr(event_list[i], attr_list[j])))
            # worksheet.write(i + 1, j, getattr(event_list[i], attr_list[j]))
            worksheet.cell(row=i + 2, column=j + 1).value = getattr(event_list[i], attr_list[j])
            # return
            # print(str(attr_list[j]) + " : " + str(getattr(event_list[i], attr_list[j])))
    # 写入磁盘文件
    workbook.save(excel_name)


log_file_list = get_file_list(root_path)
# print(log_file_list)
torque_event_list = read_file_list(root_path, log_file_list)
create_excel(excel_name, 'sheet 1', torque_event_list)
# print(get_class_attr(TorqueEventFormat))
iniParser = SqlConfigParser()
print(iniParser.host)
print(iniParser.user)
print(iniParser.password)
print(iniParser.charset)
